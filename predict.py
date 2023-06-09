import torch
from torchvision import transforms
from swin_IR import swinIR
# from unet_model import Unet
# from urdn_model import RDR_model
from unit import to_psnr, to_ssim_skimage, to_ssim, to_pearson
from PIL import Image
import matplotlib.pyplot as plt
import torch.nn as nn
import openpyxl

# wb = openpyxl.Workbook()
# ws = wb.create_sheet('celeb')  # bulid a new excel
# ws.cell(row=2, column=1).value = "mse"
# ws.cell(row=3, column=1).value = "psnr"
# ws.cell(row=4, column=1).value = "npcc"
# ws.cell(row=5, column=1).value = "ssim"
# s = 1

# Change parameter
parameter = '1.1f'

Tensor = transforms.Compose([transforms.Resize((128, 128)), transforms.ToTensor()])
model = swinIR()
device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")  # Deifne the device
model.to(device)
print("Model loaded")

Image_path = '' # Input path
label_path = '' # Groundtruth path
img = Image.open(Image_path)
if img.mode=='RGB':
    img = img.convert('L')
img = Tensor(img)
img = img / (torch.max(img))  # ImageNor
img = torch.reshape(img, (1, 1, 128, 128))

label = Image.open(label_path)
if label.mode=='RGB':
    label = label.convert('L')
label = Tensor(label)
label = label / (torch.max(label))  
label = torch.reshape(label, (1, 1, 128, 128))

model.load_state_dict(torch.load('./parameter/lr=0.0005_{}_nor'.format(parameter)))
#model.load_state_dict(torch.load('./parameter/swin_noconv_{}'.format(parameter)))
b = torch.ones(128, 128).to(device)

model.eval()
with torch.no_grad():
    img = img.to(device)
    label = label.to(device)
    output = model(img)
    output = torch.where(output < 1, output, b)
    # print(torch.max(output))
    mseloss = nn.MSELoss()
    mse = mseloss(output, label)
    psnr = to_psnr(output, label)
    npcc = to_pearson(output, label)
    print(mse)
    print(psnr)
    ssim = to_ssim_skimage(output, label)
    ssim1 = to_ssim(output, label)
    print(ssim)
    print(npcc)
    # print(ssim1)
    # s+=1
    # ws.cell(row=1, column=s).value = num
    # ws.cell(row=2, column=s).value = mse.item()
    # ws.cell(row=3, column=s).value = psnr
    # ws.cell(row=4, column=s).value = ssim[0]
    # ws.cell(row=5, column=s).value = npcc.item()
    to_pil = transforms.ToPILImage()
    output_pil = to_pil(output[0]) 
    gt = to_pil(label[0])
    plt.imshow(output_pil, cmap='gray')
    plt.axis('off')
    plt.savefig('') # save path
    plt.show()


